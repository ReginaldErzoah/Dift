from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from dift.reports.models import DiffReport

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)

TITLE_FONT = Font(size=16, bold=True, color="111827")
SECTION_FONT = Font(size=12, bold=True, color="111827")
LABEL_FONT = Font(bold=True, color="374151")

LOW_FILL = PatternFill("solid", fgColor="DCFCE7")
MEDIUM_FILL = PatternFill("solid", fgColor="FEF3C7")
HIGH_FILL = PatternFill("solid", fgColor="FEE2E2")

LOW_FONT = Font(color="166534", bold=True)
MEDIUM_FONT = Font(color="92400E", bold=True)
HIGH_FONT = Font(color="991B1B", bold=True)

LIGHT_FILL = PatternFill("solid", fgColor="F9FAFB")
SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def render_excel(report: DiffReport, output: str | None = None) -> Path:
    """
    Render and write an Excel report.

    Excel reports should always be written to a file.
    """

    output_path = Path(output or "dift_report.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    _add_summary_sheet(ws, report)
    _add_quality_sheet(wb, report)
    _add_numeric_sheet(wb, report)
    _add_outlier_sheet(wb, report)
    _add_categorical_sheet(wb, report)

    wb.save(output_path)
    return output_path


def _add_summary_sheet(ws, report: DiffReport) -> None:
    ws.append(["Dift Dataset Diff Report"])
    ws.append([])

    ws.append(["Report Metadata"])
    ws.append(["Metric", "Value"])
    ws.append(["tool", report.metadata.tool])
    ws.append(["version", report.metadata.version])
    ws.append(["report_type", report.metadata.report_type])
    ws.append(["generated_at", report.metadata.generated_at])
    ws.append(["old_source", report.metadata.old_source])
    ws.append(["new_source", report.metadata.new_source])
    ws.append(["key", report.metadata.key])
    ws.append(["threshold", report.metadata.threshold])
    ws.append(["report_format", report.metadata.report_format])
    ws.append(["template", report.metadata.template])
    ws.append(["runtime_seconds", report.metadata.runtime_seconds])

    ws.append([])
    ws.append(["Dataset Summary"])
    ws.append(["Metric", "Value"])
    ws.append(["old_rows", report.summary.old_rows])
    ws.append(["new_rows", report.summary.new_rows])
    ws.append(["row_delta", report.summary.row_delta])
    ws.append(["old_columns", report.summary.old_columns])
    ws.append(["new_columns", report.summary.new_columns])
    ws.append(["column_delta", report.summary.column_delta])
    ws.append(["risk_level", report.summary.risk_level])

    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = LIGHT_FILL

    _style_section_row(ws, 3)
    _style_header_row(ws, 4)

    _style_section_row(ws, 17)
    _style_header_row(ws, 18)

    _style_metric_column(ws)
    _style_risk_cells(ws)
    _style_delta_cells(ws)
    _apply_table_borders(ws)
    _auto_size_columns(ws)

    ws.freeze_panes = "A4"


def _add_quality_sheet(wb: Workbook, report: DiffReport) -> None:
    ws = wb.create_sheet("Quality Diff")

    ws.append(["Null Changes"])
    ws.append(
        [
            "Column",
            "Old Nulls",
            "New Nulls",
            "Old Null %",
            "New Null %",
            "Delta Null %",
            "Spike",
            "Severity",
        ]
    )

    for item in report.quality_diff.null_diffs:
        ws.append(
            [
                item.column,
                item.old_nulls,
                item.new_nulls,
                item.old_null_pct,
                item.new_null_pct,
                item.delta_null_pct,
                "Yes" if item.is_spike else "No",
                item.severity,
            ]
        )

    if not report.quality_diff.null_diffs:
        ws.append(["No null changes detected."])

    duplicate = report.quality_diff.duplicate_diff

    start_row = ws.max_row + 2
    ws.append([])
    ws.append(["Duplicate Changes"])
    ws.append(["Duplicate Metric", "Value"])
    ws.append(["Old duplicates", duplicate.old_duplicates])
    ws.append(["New duplicates", duplicate.new_duplicates])
    ws.append(["Delta duplicates", duplicate.delta_duplicates])
    ws.append(["Old duplicate %", duplicate.old_duplicate_pct])
    ws.append(["New duplicate %", duplicate.new_duplicate_pct])
    ws.append(["Delta duplicate %", duplicate.delta_duplicate_pct])
    ws.append(["Duplicate basis", duplicate.duplicate_basis])
    ws.append(["Spike", "Yes" if duplicate.is_spike else "No"])
    ws.append(["Severity", duplicate.severity])

    _style_section_row(ws, 1)
    _style_header_row(ws, 2)
    _style_section_row(ws, start_row)
    _style_header_row(ws, start_row + 1)

    _style_severity_cells(ws)
    _style_spike_cells(ws)
    _style_delta_cells(ws)
    _apply_table_borders(ws)
    _auto_size_columns(ws)

    ws.freeze_panes = "A3"


def _add_numeric_sheet(wb: Workbook, report: DiffReport) -> None:
    ws = wb.create_sheet("Numeric Drift")

    ws.append(
        [
            "Column",
            "Old Mean",
            "New Mean",
            "Delta Mean",
            "Mean Shift %",
            "Old Std",
            "New Std",
            "Delta Std",
            "Std Shift %",
            "Delta Range",
            "Range Shift %",
            "Threshold",
            "Drifted",
            "Severity",
        ]
    )

    for item in report.numeric_diff:
        ws.append(
            [
                item.column,
                item.old_mean,
                item.new_mean,
                item.delta_mean,
                item.mean_shift_pct,
                item.old_std,
                item.new_std,
                item.delta_std,
                item.std_shift_pct,
                item.delta_range,
                item.range_shift_pct,
                item.drift_threshold,
                "Yes" if item.is_drifted else "No",
                item.severity,
            ]
        )

    if not report.numeric_diff:
        ws.append(["No numeric drift detected."])

    _style_header_row(ws, 1)
    _style_severity_cells(ws)
    _style_spike_cells(ws)
    _highlight_yes_rows(ws, "Drifted")
    _apply_table_borders(ws)
    _auto_size_columns(ws)

    ws.freeze_panes = "A2"


def _add_outlier_sheet(wb: Workbook, report: DiffReport) -> None:
    ws = wb.create_sheet("Outlier Diff")

    ws.append(
        [
            "Column",
            "Method",
            "Old Outliers",
            "New Outliers",
            "Delta Outliers",
            "Old Outlier %",
            "New Outlier %",
            "Delta Outlier %",
            "Lower Bound",
            "Upper Bound",
            "Spike",
            "Severity",
        ]
    )

    for item in report.outlier_diff:
        ws.append(
            [
                item.column,
                item.method,
                item.old_outliers,
                item.new_outliers,
                item.delta_outliers,
                item.old_outlier_pct,
                item.new_outlier_pct,
                item.delta_outlier_pct,
                item.lower_bound,
                item.upper_bound,
                "Yes" if item.is_spike else "No",
                item.severity,
            ]
        )

    if not report.outlier_diff:
        ws.append(["No outlier changes detected."])

    _style_header_row(ws, 1)
    _style_severity_cells(ws)
    _style_spike_cells(ws)
    _highlight_yes_rows(ws, "Spike")
    _apply_table_borders(ws)
    _auto_size_columns(ws)

    ws.freeze_panes = "A2"


def _add_categorical_sheet(wb: Workbook, report: DiffReport) -> None:
    ws = wb.create_sheet("Categorical Diff")

    ws.append(
        [
            "Column",
            "Values Added",
            "Values Removed",
            "Frequency Shifts",
            "Max Frequency Shift",
            "Shifted",
            "Severity",
        ]
    )

    for item in report.categorical_diff:
        ws.append(
            [
                item.column,
                ", ".join(item.values_added),
                ", ".join(item.values_removed),
                _format_frequency_shifts(item.frequency_shifts),
                item.max_frequency_shift,
                "Yes" if item.is_shifted else "No",
                item.severity,
            ]
        )

    if not report.categorical_diff:
        ws.append(["No categorical changes detected."])

    _style_header_row(ws, 1)
    _style_severity_cells(ws)
    _style_spike_cells(ws)
    _highlight_yes_rows(ws, "Shifted")
    _apply_table_borders(ws)
    _auto_size_columns(ws)

    ws.freeze_panes = "A2"


def _format_frequency_shifts(shifts: dict[str, float]) -> str:
    if not shifts:
        return ""

    return ", ".join(f"{value}: {shift:.2%}" for value, shift in shifts.items())


def _style_header_row(ws, row_number: int) -> None:
    for cell in ws[row_number]:
        if cell.value is None:
            continue

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_section_row(ws, row_number: int) -> None:
    for cell in ws[row_number]:
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER


def _style_metric_column(ws) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        first_cell = row[0]
        if first_cell.value and first_cell.value not in {
            "Metric",
            "Value",
            "Report Metadata",
            "Dataset Summary",
            "Dift Dataset Diff Report",
        }:
            first_cell.font = LABEL_FONT


def _style_risk_cells(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            value = _normalize(cell.value)

            if value in {"low", "medium", "high"}:
                _apply_severity_style(cell, value)


def _style_severity_cells(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            value = _normalize(cell.value)

            if value in {"low", "medium", "high"}:
                _apply_severity_style(cell, value)


def _style_spike_cells(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            value = _normalize(cell.value)

            if value == "yes":
                cell.fill = HIGH_FILL
                cell.font = HIGH_FONT

            elif value == "no":
                cell.fill = LOW_FILL
                cell.font = LOW_FONT


def _style_delta_cells(ws) -> None:
    for row in ws.iter_rows():
        label = _normalize(row[0].value)

        if "delta" not in label:
            continue

        for cell in row[1:]:
            if isinstance(cell.value, int | float) and cell.value != 0:
                cell.fill = MEDIUM_FILL
                cell.font = MEDIUM_FONT


def _highlight_yes_rows(ws, column_name: str) -> None:
    target_col = None

    for cell in ws[1]:
        if _normalize(cell.value) == _normalize(column_name):
            target_col = cell.column
            break

    if target_col is None:
        return

    for row_number in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_number, column=target_col)

        if _normalize(cell.value) == "yes":
            for row_cell in ws[row_number]:
                if row_cell.fill == PatternFill(fill_type=None):
                    row_cell.fill = HIGH_FILL


def _apply_severity_style(cell, value: str) -> None:
    if value == "high":
        cell.fill = HIGH_FILL
        cell.font = HIGH_FONT
    elif value == "medium":
        cell.fill = MEDIUM_FILL
        cell.font = MEDIUM_FONT
    elif value == "low":
        cell.fill = LOW_FILL
        cell.font = LOW_FONT


def _apply_table_borders(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            cell.border = THIN_BORDER
            cell.alignment = LEFT


def _auto_size_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 4, 60)


def _normalize(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip().lower()