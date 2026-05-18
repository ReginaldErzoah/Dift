from openpyxl import load_workbook

from dift.reports.excel_report import render_excel
from dift.reports.models import (
    CategoricalDiff,
    DiffReport,
    DuplicateDiff,
    NullDiff,
    NumericDiff,
    OutlierDiff,
    QualityDiff,
    RowDiff,
    SchemaDiff,
    Summary,
)


def _sample_report() -> DiffReport:
    return DiffReport(
        summary=Summary(
            old_rows=10,
            new_rows=12,
            row_delta=2,
            old_columns=4,
            new_columns=5,
            column_delta=1,
            risk_level="high",
        ),
        schema_diff=SchemaDiff(columns_added=["email"]),
        row_diff=RowDiff(
            key="customer_id",
            added_rows=2,
            removed_rows=0,
            changed_rows=1,
            unchanged_rows=9,
            compared_columns=["name", "email", "revenue"],
        ),
        quality_diff=QualityDiff(
            null_diffs=[
                NullDiff(
                    column="email",
                    old_nulls=0,
                    new_nulls=3,
                    old_null_pct=0.0,
                    new_null_pct=25.0,
                    delta_null_pct=25.0,
                    is_spike=True,
                    severity="high",
                )
            ],
            duplicate_diff=DuplicateDiff(
                old_duplicates=0,
                new_duplicates=2,
                delta_duplicates=2,
                old_duplicate_pct=0.0,
                new_duplicate_pct=16.67,
                delta_duplicate_pct=16.67,
                duplicate_basis="all_columns",
                is_spike=True,
                severity="medium",
            ),
        ),
        numeric_diff=[
            NumericDiff(
                column="revenue",
                old_mean=100.0,
                new_mean=250.0,
                delta_mean=150.0,
                mean_shift_pct=1.5,
                old_std=10.0,
                new_std=40.0,
                delta_std=30.0,
                std_shift_pct=3.0,
                delta_range=100.0,
                range_shift_pct=1.0,
                drift_threshold=0.1,
                is_drifted=True,
                severity="high",
            )
        ],
        outlier_diff=[
            OutlierDiff(
                column="revenue",
                method="iqr",
                old_outliers=0,
                new_outliers=2,
                delta_outliers=2,
                old_outlier_pct=0.0,
                new_outlier_pct=16.67,
                delta_outlier_pct=16.67,
                lower_bound=10.0,
                upper_bound=300.0,
                is_spike=True,
                severity="high",
            )
        ],
        categorical_diff=[
            CategoricalDiff(
                column="segment",
                values_added=["enterprise"],
                values_removed=[],
                old_top_values={"retail": 8},
                new_top_values={"enterprise": 6},
                frequency_shifts={"enterprise": 0.6},
                max_frequency_shift=0.6,
                is_shifted=True,
                severity="high",
            )
        ],
    )


def test_render_excel_writes_file(tmp_path):
    output_path = tmp_path / "report.xlsx"
    report = _sample_report()

    result = render_excel(report, output=str(output_path))

    assert result == output_path
    assert output_path.exists()


def test_excel_report_contains_expected_sheets(tmp_path):
    output_path = tmp_path / "report.xlsx"
    report = _sample_report()

    render_excel(report, output=str(output_path))

    wb = load_workbook(output_path)

    assert "Summary" in wb.sheetnames
    assert "Quality Diff" in wb.sheetnames
    assert "Numeric Drift" in wb.sheetnames
    assert "Outlier Diff" in wb.sheetnames
    assert "Categorical Diff" in wb.sheetnames


def test_summary_sheet_layout_and_risk_highlight(tmp_path):
    output_path = tmp_path / "report.xlsx"
    report = _sample_report()

    render_excel(report, output=str(output_path))

    wb = load_workbook(output_path)
    ws = wb["Summary"]

    assert ws["A1"].value == "Dift Dataset Diff Report"
    assert ws.freeze_panes == "A4"

    values = [cell.value for row in ws.iter_rows() for cell in row]

    assert "Report Metadata" in values
    assert "Dataset Summary" in values
    assert "risk_level" in values
    assert "high" in values

    risk_cell = None
    for row in ws.iter_rows():
        if row[0].value == "risk_level":
            risk_cell = row[1]
            break

    assert risk_cell is not None
    assert risk_cell.fill.fill_type == "solid"
    assert risk_cell.font.bold is True


def test_header_rows_are_styled(tmp_path):
    output_path = tmp_path / "report.xlsx"
    report = _sample_report()

    render_excel(report, output=str(output_path))

    wb = load_workbook(output_path)

    for sheet_name in ["Numeric Drift", "Outlier Diff", "Categorical Diff"]:
        ws = wb[sheet_name]
        assert ws["A1"].font.bold is True
        assert ws["A1"].fill.fill_type == "solid"
        assert ws.freeze_panes == "A2"


def test_severity_cells_are_color_coded(tmp_path):
    output_path = tmp_path / "report.xlsx"
    report = _sample_report()

    render_excel(report, output=str(output_path))

    wb = load_workbook(output_path)
    ws = wb["Numeric Drift"]

    severity_cell = ws["N2"]

    assert severity_cell.value == "high"
    assert severity_cell.fill.fill_type == "solid"
    assert severity_cell.font.bold is True


def test_spike_and_drift_cells_are_highlighted(tmp_path):
    output_path = tmp_path / "report.xlsx"
    report = _sample_report()

    render_excel(report, output=str(output_path))

    wb = load_workbook(output_path)

    numeric_ws = wb["Numeric Drift"]
    outlier_ws = wb["Outlier Diff"]
    categorical_ws = wb["Categorical Diff"]

    assert numeric_ws["M2"].value == "Yes"
    assert numeric_ws["M2"].fill.fill_type == "solid"

    assert outlier_ws["K2"].value == "Yes"
    assert outlier_ws["K2"].fill.fill_type == "solid"

    assert categorical_ws["F2"].value == "Yes"
    assert categorical_ws["F2"].fill.fill_type == "solid"


def test_column_widths_are_auto_adjusted(tmp_path):
    output_path = tmp_path / "report.xlsx"
    report = _sample_report()

    render_excel(report, output=str(output_path))

    wb = load_workbook(output_path)
    ws = wb["Summary"]

    assert ws.column_dimensions["A"].width is not None
    assert ws.column_dimensions["A"].width > 10